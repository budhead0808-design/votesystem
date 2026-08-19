import openpyxl
import re
from db import init_db, get_db, generate_voter_code

def import_school_staff():
    init_db()
    conn = get_db()
    cursor = conn.cursor()

    file_path = '115學年度教職員工名單_代理教師標註.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['完整名單'] if '完整名單' in wb.sheetnames else wb.active

    rows = [[str(c).strip() if c is not None else '' for c in r] for r in sheet.iter_rows(values_only=True) if any(r)]
    
    header_idx = None
    name_col = 4
    role_col = 1
    type_col = 5

    for idx, r in enumerate(rows):
        if any('姓名' in str(c) for c in r):
            header_idx = idx
            for c_i, cell in enumerate(r):
                if '姓名' in cell:
                    name_col = c_i
                elif '類別' in cell:
                    role_col = c_i
                elif '身分類別' in cell:
                    type_col = c_i
            break

    start_row = (header_idx + 1) if header_idx is not None else 3
    teachers_added = 0
    teachers_updated = 0

    for row in rows[start_row:]:
        if len(row) <= name_col:
            continue
        name = row[name_col]
        # 清理姓名空格 (例: "王  煜" -> "王煜")
        name = re.sub(r'\s+', '', name)
        
        if not name or name in ['姓名', '原始內容', '借調4', '借調1', '借調2', '借調3'] or len(name) < 2:
            continue

        role_str = row[role_col] if len(row) > role_col else ''
        type_str = row[type_col] if len(row) > type_col else ''

        is_admin = 1 if ('行政' in role_str or '主任' in role_str or '組長' in role_str) else 0

        cursor.execute("SELECT id FROM teachers WHERE name = ?", (name,))
        existing = cursor.fetchone()
        if not existing:
            vcode = generate_voter_code()
            cursor.execute("""
                INSERT INTO teachers (name, gender, department, is_admin, voter_code)
                VALUES (?, ?, ?, ?, ?)
            """, (name, '男', '一般', is_admin, vcode))
            teachers_added += 1
        else:
            cursor.execute("UPDATE teachers SET is_admin = ? WHERE name = ?", (is_admin, name))
            teachers_updated += 1

    conn.commit()

    # 統計目前全校投票權人總數
    cursor.execute("SELECT COUNT(*) FROM teachers")
    total_teachers = cursor.fetchone()[0]
    conn.close()

    print(f"Import finished: Added {teachers_added} new staff, updated {teachers_updated}. Total school voting roster: {total_teachers} members.")

if __name__ == '__main__':
    import_school_staff()
