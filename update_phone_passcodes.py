import openpyxl
import re
from db import init_db, get_db

def update_passcodes():
    init_db()
    conn = get_db()
    cursor = conn.cursor()

    excel_path = '115學年度教職員工名單_代理教師標註_已填手機.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb['完整名單'] if '完整名單' in wb.sheetnames else (wb['名冊'] if '名冊' in wb.sheetnames else wb.active)

    rows = [[str(c).strip() if c is not None else '' for c in r] for r in sheet.iter_rows(values_only=True) if any(r)]
    
    header_idx = None
    name_col = 4
    phone_col = 8

    for idx, r in enumerate(rows):
        if any('姓名' in str(c) for c in r):
            header_idx = idx
            for c_i, cell in enumerate(r):
                if '姓名' in cell:
                    name_col = c_i
                elif '手機' in cell:
                    phone_col = c_i
            break

    start_row = (header_idx + 1) if header_idx is not None else 3

    phone_map = {}
    for row in rows[start_row:]:
        if len(row) <= name_col:
            continue
        name = re.sub(r'\s+', '', row[name_col])
        if not name or name in ['姓名', '原始內容'] or len(name) < 2:
            continue

        phone_raw = row[phone_col] if len(row) > phone_col else ''
        digits = re.sub(r'\D', '', phone_raw)
        if len(digits) >= 4:
            passcode = digits[-4:]
        else:
            passcode = '5313'

        phone_map[name] = passcode

    # 更新資料庫全校教師之投票密碼 voter_code
    cursor.execute("SELECT id, name FROM teachers")
    db_teachers = cursor.fetchall()

    updated_phone_count = 0
    default_count = 0

    for t in db_teachers:
        t_id = t['id']
        t_name = re.sub(r'\s+', '', t['name'])
        code = phone_map.get(t_name, '5313')

        if code == '5313':
            default_count += 1
        else:
            updated_phone_count += 1

        cursor.execute("UPDATE teachers SET voter_code = ? WHERE id = ?", (code, t_id))

    conn.commit()
    conn.close()

    print(f"Passcode update finished! {updated_phone_count} teachers set to phone last 4 digits, {default_count} teachers set to default 5313.")

if __name__ == '__main__':
    update_passcodes()
