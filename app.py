import os
import io
import re
from flask import Flask, render_template, request, jsonify, send_file
from db import get_db, init_db, generate_voter_code
from gender_rules import calculate_election_results
from word_parser import parse_word_docx
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.secret_key = 'school_voting_system_secret_key'

# 初始化資料庫
init_db()

# --- 頁面路由 ---

@app.route('/')
def index():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM committees ORDER BY id DESC")
    committees = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT COUNT(*) FROM teachers")
    teacher_count = cursor.fetchone()[0]
    conn.close()

    active_count = sum(1 for c in committees if c['status'] == 'active')
    closed_count = sum(1 for c in committees if c['status'] == 'closed')

    return render_template('index.html', 
                           committees=committees, 
                           active_count=active_count, 
                           closed_count=closed_count,
                           teacher_count=teacher_count)

@app.route('/committee/new')
def committee_new():
    return render_template('committee_form.html', committee=None)

@app.route('/committee/edit/<int:committee_id>')
def committee_edit(committee_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM committees WHERE id = ?", (committee_id,))
    committee = cursor.fetchone()
    conn.close()
    if not committee:
        return "投票項目不存在", 404
    return render_template('committee_form.html', committee=dict(committee))

@app.route('/roster')
def roster():
    return render_template('roster.html')

@app.route('/vote/<int:committee_id>')
def vote(committee_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM committees WHERE id = ?", (committee_id,))
    committee = cursor.fetchone()
    cursor.execute("SELECT * FROM teachers ORDER BY id ASC")
    teachers = [dict(r) for r in cursor.fetchall()]
    conn.close()
    if not committee:
        return "投票項目不存在", 404
    return render_template('vote.html', committee=dict(committee), teachers=teachers)

@app.route('/manual_count/<int:committee_id>')
def manual_count(committee_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM committees WHERE id = ?", (committee_id,))
    committee = cursor.fetchone()
    conn.close()
    if not committee:
        return "投票項目不存在", 404
    return render_template('manual_count.html', committee=dict(committee))

@app.route('/result/<int:committee_id>')
def result(committee_id):
    res = calculate_election_results(committee_id)
    if not res:
        return "找不到開票結果", 404
    return render_template('result.html', result=res)

@app.route('/print_ballot/<int:committee_id>')
def print_ballot(committee_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM committees WHERE id = ?", (committee_id,))
    committee = cursor.fetchone()
    cursor.execute("SELECT * FROM candidates WHERE committee_id = ? ORDER BY candidate_number ASC", (committee_id,))
    candidates = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return render_template('print_ballot.html', committee=dict(committee), candidates=candidates)

@app.route('/print_passcodes')
def print_passcodes():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM teachers ORDER BY id ASC")
    teachers = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return render_template('roster.html', teachers=teachers, print_mode=True)


# --- API 路由 ---

@app.route('/api/committees', methods=['GET'])
def get_committees():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM committees ORDER BY id DESC")
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/committee', methods=['POST'])
def create_committee():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO committees (
            title, description, seats_count, alternate_count, max_votes_per_ballot,
            auth_mode, show_gender, show_admin, gender_rule_type, min_male_count, min_female_count,
            identity_rule_type, min_non_admin_count, status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.get('title'),
        data.get('description', ''),
        data.get('seats_count', 5),
        data.get('alternate_count', 0),
        data.get('max_votes_per_ballot', 4),
        data.get('auth_mode', 'passcode'),
        data.get('show_gender', 1),
        data.get('show_admin', 1),
        data.get('gender_rule_type', 'none'),
        data.get('min_male_count', 0),
        data.get('min_female_count', 0),
        data.get('identity_rule_type', 'none'),
        data.get('min_non_admin_count', 0),
        data.get('status', 'active')
    ))
    cid = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': cid})

@app.route('/api/committee/<int:committee_id>', methods=['PUT', 'DELETE'])
def update_delete_committee(committee_id):
    conn = get_db()
    cursor = conn.cursor()
    if request.method == 'DELETE':
        cursor.execute("DELETE FROM committees WHERE id = ?", (committee_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    data = request.json
    cursor.execute("""
        UPDATE committees SET 
            title=?, description=?, seats_count=?, alternate_count=?, max_votes_per_ballot=?,
            auth_mode=?, show_gender=?, show_admin=?, gender_rule_type=?, min_male_count=?, min_female_count=?,
            identity_rule_type=?, min_non_admin_count=?, status=?
        WHERE id = ?
    """, (
        data.get('title'), data.get('description'), data.get('seats_count'), data.get('alternate_count'),
        data.get('max_votes_per_ballot'), data.get('auth_mode', 'passcode'), data.get('show_gender', 1),
        data.get('show_admin', 1), data.get('gender_rule_type'), data.get('min_male_count'),
        data.get('min_female_count'), data.get('identity_rule_type'), data.get('min_non_admin_count'),
        data.get('status'), committee_id
    ))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/candidate/<int:candidate_id>', methods=['DELETE'])
def delete_candidate(candidate_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM candidates WHERE id = ?", (candidate_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/teachers', methods=['GET', 'POST'])
def handle_teachers():
    conn = get_db()
    cursor = conn.cursor()
    if request.method == 'GET':
        cursor.execute("SELECT * FROM teachers ORDER BY id ASC")
        teachers = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify(teachers)
    
    data = request.json
    voter_code = data.get('voter_code') or generate_voter_code()
    cursor.execute("""
        INSERT INTO teachers (name, gender, department, is_admin, voter_code)
        VALUES (?, ?, ?, ?, ?)
    """, (data.get('name'), data.get('gender', '男'), '一般', data.get('is_admin', 0), voter_code))
    tid = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': tid})

@app.route('/api/committee/<int:committee_id>/batch_add_candidates', methods=['POST'])
def batch_add_candidates(committee_id):
    data = request.json
    names_text = data.get('names_text', '')
    
    raw_lines = re.split(r'[\r\n]+', names_text)
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT MAX(candidate_number) FROM candidates WHERE committee_id = ?", (committee_id,))
    max_num = cursor.fetchone()[0] or 0

    added_count = 0
    current_num = max_num + 1

    for line in raw_lines:
        line_clean = line.strip()
        if not line_clean:
            continue

        match = re.search(r'(?:(\d+)[\.\s、:]*)?([\u4e00-\u9fa5]{2,4})', line_clean)
        if match:
            num_str, name = match.groups()
            num = int(num_str) if num_str else current_num
            
            cursor.execute("SELECT id FROM candidates WHERE committee_id = ? AND name = ?", (committee_id, name))
            existing_c = cursor.fetchone()
            if existing_c:
                cursor.execute("UPDATE candidates SET candidate_number = ? WHERE id = ?", (num, existing_c['id']))
            else:
                cursor.execute("""
                    INSERT INTO candidates (committee_id, candidate_number, name, gender, department, is_admin)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (committee_id, num, name, '男', '一般', 0))
                added_count += 1
                current_num = max(current_num + 1, num + 1)

            cursor.execute("SELECT id FROM teachers WHERE name = ?", (name,))
            if not cursor.fetchone():
                vcode = '5313'
                cursor.execute("INSERT INTO teachers (name, gender, department, is_admin, voter_code) VALUES (?, ?, ?, ?, ?)",
                               (name, '男', '一般', 0, vcode))

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'added_count': added_count})

@app.route('/api/batch_add_voters', methods=['POST'])
def batch_add_voters():
    data = request.json
    names_text = data.get('names_text', '')
    raw_names = re.split(r'[\r\n,，;\s]+', names_text)
    names = [n.strip() for n in raw_names if len(n.strip()) >= 2]

    if not names:
        return jsonify({'success': False, 'message': '請輸入有效的姓名'}), 400

    conn = get_db()
    cursor = conn.cursor()
    added_count = 0
    for name in names:
        cursor.execute("SELECT id FROM teachers WHERE name = ?", (name,))
        if not cursor.fetchone():
            vcode = '5313'
            cursor.execute("INSERT INTO teachers (name, gender, department, is_admin, voter_code) VALUES (?, ?, ?, ?, ?)",
                           (name, '男', '一般', 0, vcode))
            added_count += 1

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'added_count': added_count, 'total_submitted': len(names)})

@app.route('/api/teachers/<int:teacher_id>', methods=['DELETE'])
def delete_teacher(teacher_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM teachers WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/generate_voter_codes', methods=['POST'])
def generate_codes():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM teachers")
    teachers = cursor.fetchall()
    for t in teachers:
        code = generate_voter_code()
        cursor.execute("UPDATE teachers SET voter_code = ? WHERE id = ?", (code, t['id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'count': len(teachers)})

@app.route('/api/upload_word', methods=['POST'])
def upload_word():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未上傳檔案'}), 400
    file = request.files['file']
    filename = file.filename.lower()

    parsed_candidates = []
    try:
        if filename.endswith('.docx'):
            parsed_candidates = parse_word_docx(file)
        elif filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv'):
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            
            rows = [[str(c).strip() if c is not None else '' for c in r] for r in sheet.iter_rows(values_only=True) if any(r)]
            if not rows:
                return jsonify({'success': True, 'candidates': []})

            header_idx = None
            num_col = None
            name_col = None

            for idx, row in enumerate(rows):
                for col_i, cell in enumerate(row):
                    if any(k in cell for k in ['號', '編號', '號次', '序號', 'No']):
                        num_col = col_i
                    elif any(k in cell for k in ['姓名', '候選人', '教師姓名', '名單']):
                        name_col = col_i

                if name_col is not None:
                    header_idx = idx
                    break

            start_row = (header_idx + 1) if header_idx is not None else 0
            current_num = 1

            for row in rows[start_row:]:
                if not any(row):
                    continue

                name = ""
                num = current_num

                if name_col is not None and name_col < len(row):
                    name = row[name_col]
                else:
                    for c in row:
                        clean_c = re.sub(r'[\d\.\s]', '', c)
                        if 2 <= len(clean_c) <= 4 and clean_c not in ['號次', '姓名', '編號', '備註', '名單總數']:
                            name = clean_c
                            break

                if not name or any(k in name for k in ['名單總數', '備註', '說明', '初選名單']):
                    continue

                if num_col is not None and num_col < len(row):
                    try:
                        num = int(re.sub(r'\D', '', row[num_col]))
                    except:
                        num = current_num
                else:
                    if row[0].isdigit():
                        num = int(row[0])

                parsed_candidates.append({
                    'candidate_number': num,
                    'name': name,
                    'gender': '男',
                    'is_admin': 0
                })
                current_num += 1

        return jsonify({'success': True, 'candidates': parsed_candidates})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Excel 檔案解析出錯: {str(e)}'}), 500

@app.route('/api/committee/<int:committee_id>/import_candidates', methods=['POST'])
def import_candidates(committee_id):
    data = request.json
    candidates = data.get('candidates', [])
    target_type = data.get('target_type', 'both')

    if not candidates:
        return jsonify({'success': False, 'message': '沒有可匯入的名單'}), 400

    conn = get_db()
    cursor = conn.cursor()
    
    if target_type == 'both':
        # 智慧 UPSERT 模式：保留既有候選人與其歷史得票紀錄！
        count = 0
        for idx, c in enumerate(candidates, start=1):
            num = c.get('candidate_number') or idx
            name = c.get('name')
            if not name:
                continue

            gender = c.get('gender') if c.get('gender') in ['男', '女'] else '男'
            is_admin = c.get('is_admin', 0)

            cursor.execute("SELECT id FROM candidates WHERE committee_id = ? AND name = ?", (committee_id, name))
            existing_c = cursor.fetchone()

            if existing_c:
                cursor.execute("""
                    UPDATE candidates SET candidate_number = ?, gender = ?, is_admin = ? WHERE id = ?
                """, (num, gender, is_admin, existing_c['id']))
            else:
                cursor.execute("""
                    INSERT INTO candidates (committee_id, candidate_number, name, gender, department, is_admin)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (committee_id, num, name, gender, '一般', is_admin))
                count += 1

            cursor.execute("SELECT id FROM teachers WHERE name = ?", (name,))
            if not cursor.fetchone():
                vcode = '5313'
                cursor.execute("INSERT INTO teachers (name, gender, department, is_admin, voter_code) VALUES (?, ?, ?, ?, ?)",
                               (name, gender, '一般', is_admin, vcode))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'imported_count': count, 'type': 'both'})

    else:
        count = 0
        for c in candidates:
            name = c.get('name')
            if not name:
                continue
            cursor.execute("SELECT id FROM teachers WHERE name = ?", (name,))
            if not cursor.fetchone():
                vcode = '5313'
                cursor.execute("INSERT INTO teachers (name, gender, department, is_admin, voter_code) VALUES (?, ?, ?, ?, ?)",
                               (name, '男', '一般', 0, vcode))
                count += 1
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'imported_count': count, 'type': 'voters_only'})

@app.route('/api/committee/<int:committee_id>/candidates', methods=['GET'])
def get_committee_candidates(committee_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM candidates WHERE committee_id = ? ORDER BY candidate_number ASC", (committee_id,))
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/committee/<int:committee_id>/voted_teachers', methods=['GET'])
def get_voted_teachers(committee_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT teacher_id FROM voter_logs WHERE committee_id = ?", (committee_id,))
    voted_ids = [r['teacher_id'] for r in cursor.fetchall()]
    conn.close()
    return jsonify(voted_ids)

@app.route('/api/committee/<int:committee_id>/candidates_with_votes', methods=['GET'])
def get_candidates_with_votes(committee_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT c.*, 
            COALESCE(b.online_votes, 0) + COALESCE(m.manual_votes, 0) AS total_votes
        FROM candidates c
        LEFT JOIN (
            SELECT candidate_id, COUNT(*) as online_votes 
            FROM ballots WHERE committee_id = ? GROUP BY candidate_id
        ) b ON c.id = b.candidate_id
        LEFT JOIN (
            SELECT candidate_id, SUM(votes_count) as manual_votes 
            FROM manual_votes WHERE committee_id = ? GROUP BY candidate_id
        ) m ON c.id = m.candidate_id
        WHERE c.committee_id = ?
        ORDER BY c.candidate_number ASC
    """, (committee_id, committee_id, committee_id))
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/voter_auth', methods=['POST'])
def voter_auth():
    data = request.json
    cid = data.get('committee_id')
    code = data.get('voter_code', '').strip().upper()

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM teachers WHERE UPPER(voter_code) = ?", (code,))
    teacher = cursor.fetchone()
    if not teacher:
        conn.close()
        return jsonify({'success': False, 'message': '無效的投票密碼'}), 400

    teacher = dict(teacher)
    cursor.execute("SELECT id FROM voter_logs WHERE committee_id = ? AND teacher_id = ?", (cid, teacher['id']))
    if cursor.fetchone():
        conn.close()
        return jsonify({'success': False, 'message': '您在此投票項目中已經完成過投票，無法重複投票'}), 400

    conn.close()
    return jsonify({'success': True, 'teacher': teacher})

@app.route('/api/vote', methods=['POST'])
def submit_vote():
    data = request.json
    cid = data.get('committee_id')
    tid = data.get('teacher_id')
    candidate_ids = data.get('selected_candidate_ids', [])

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT max_votes_per_ballot, auth_mode FROM committees WHERE id = ?", (cid,))
    c_info = cursor.fetchone()
    if not c_info:
        conn.close()
        return jsonify({'success': False, 'message': '找不到該投票項目'}), 400

    max_limit = c_info['max_votes_per_ballot']
    auth_mode = c_info['auth_mode']

    if len(candidate_ids) > max_limit:
        conn.close()
        return jsonify({'success': False, 'message': f'圈選人數超過上限 ({max_limit} 人)'}), 400

    if tid and auth_mode != 'public':
        cursor.execute("SELECT id FROM voter_logs WHERE committee_id = ? AND teacher_id = ?", (cid, tid))
        if cursor.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': '該位老師已經領票/投票過了，無法重複投票'}), 400

    for cand_id in candidate_ids:
        cursor.execute("INSERT INTO ballots (committee_id, candidate_id) VALUES (?, ?)", (cid, cand_id))

    if tid:
        cursor.execute("INSERT INTO voter_logs (committee_id, teacher_id) VALUES (?, ?)", (cid, tid))

    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/manual_vote', methods=['POST'])
def manual_vote():
    data = request.json
    cid = data.get('committee_id')
    cand_id = data.get('candidate_id')
    count = data.get('votes_count', 1)

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO manual_votes (committee_id, candidate_id, votes_count) VALUES (?, ?, ?)",
                   (cid, cand_id, count))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/export_excel/<int:committee_id>')
def export_excel(committee_id):
    res = calculate_election_results(committee_id)
    if not res:
        return "找不到資料", 404

    c_info = res['committee']
    show_gender = c_info.get('show_gender', 1)
    show_admin = c_info.get('show_admin', 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "開票結果與統計報表"

    title_font = Font(name="微軟正黑體", size=16, bold=True, color="1F4E78")
    header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    ws.append([c_info['title'] + " - 完整開票報告書"])
    ws.merge_cells("A1:E1")
    ws['A1'].font = title_font

    ws.append([])

    headers = ["當選類別", "號次", "當選人姓名"]
    if show_gender:
        headers.append("性別")
    if show_admin:
        headers.append("行政身分")
    headers.append("總得票數")

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    row_idx = 4
    for w in res['winners']:
        row_data = ["正取當選", w['candidate_number'], w['name']]
        if show_gender:
            row_data.append(w['gender'])
        if show_admin:
            row_data.append('兼行政' if w['is_admin'] else '未兼行政')
        row_data.append(w['total_votes'])
        ws.append(row_data)
        row_idx += 1

    for a in res['alternates']:
        row_data = ["候補名單", a['candidate_number'], a['name']]
        if show_gender:
            row_data.append(a['gender'])
        if show_admin:
            row_data.append('兼行政' if a['is_admin'] else '未兼行政')
        row_data.append(a['total_votes'])
        ws.append(row_data)
        row_idx += 1

    ws.append([])
    ws.append(["【開票與運算歷程】"])
    for log in res['logs']:
        ws.append([log])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"election_result_{committee_id}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == '__main__':
    print("啟動 學校教師委員會投票與智慧計票系統...")
    app.run(host='0.0.0.0', port=5000, debug=True)
